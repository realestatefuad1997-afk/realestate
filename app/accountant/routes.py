from flask import Blueprint, render_template, abort, request, redirect, url_for, flash
from flask_login import login_required, current_user
from flask_babel import gettext as _
from ..extensions import db
from ..models import (
    Payment,
    Invoice,
    Account,
    JournalEntry,
    JournalLine,
    Expense,
    Contract,
    User,
    Property,
    Apartment,
    MaintenanceRequest,
    Complaint,
)
from flask import current_app, send_file
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from openpyxl import Workbook
import io
import os
from datetime import date, timedelta
from sqlalchemy import text


accountant_bp = Blueprint("accountant", __name__)


def accountant_required(func):
    from functools import wraps

    @wraps(func)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or not (current_user.is_accountant or current_user.is_admin):
            return abort(403)
        return func(*args, **kwargs)

    return wrapper


@accountant_bp.route("/")
@login_required
@accountant_required
def dashboard():
    """Accountant dashboard with KPIs, monthly chart, and alerts."""
    # Payments listing (all)
    payments = Payment.query.order_by(Payment.due_date.asc()).all()

    # KPI cards
    total_income = (
        db.session.query(db.func.coalesce(db.func.sum(Payment.amount), 0))
        .filter(Payment.status == "paid")
        .scalar()
        or 0
    )

    # Expenses total (if table exists)
    total_expenses = 0
    try:
        total_expenses = (
            db.session.query(db.func.coalesce(db.func.sum(Expense.amount), 0)).scalar() or 0
        )
    except Exception:
        total_expenses = 0

    net_profit = float(total_income) - float(total_expenses)

    # Unpaid count
    unpaid_count = db.session.query(db.func.count()).select_from(Payment).filter(Payment.status != "paid").scalar() or 0

    # Alerts: due soon (within 7 days) and overdue
    today = date.today()
    in_7 = today + timedelta(days=7)
    due_soon = (
        db.session.query(Payment)
        .filter(Payment.status != "paid", Payment.due_date >= today, Payment.due_date <= in_7)
        .order_by(Payment.due_date.asc())
        .all()
    )
    overdue = (
        db.session.query(Payment)
        .filter(Payment.status != "paid", Payment.due_date < today)
        .order_by(Payment.due_date.asc())
        .all()
    )

    # Monthly income/expenses/profit for last 12 months
    def add_months(year: int, month: int, delta: int) -> tuple[int, int]:
        total = year * 12 + (month - 1) + delta
        y = total // 12
        m = (total % 12) + 1
        return y, m

    def first_of_month(y: int, m: int) -> date:
        return date(y, m, 1)

    def next_month_start(y: int, m: int) -> date:
        y2, m2 = add_months(y, m, 1)
        return date(y2, m2, 1)

    y0, m0 = today.year, today.month
    month_ranges: list[tuple[date, date]] = []
    month_labels: list[str] = []
    for k in range(11, -1, -1):
        yk, mk = add_months(y0, m0, -k)
        start = first_of_month(yk, mk)
        end = next_month_start(yk, mk)
        month_ranges.append((start, end))
        month_labels.append(f"{yk:04d}-{mk:02d}")

    monthly_income: list[float] = []
    for start, end in month_ranges:
        total = (
            db.session.query(db.func.coalesce(db.func.sum(Payment.amount), 0))
            .filter(
                Payment.status == "paid",
                db.or_(
                    db.and_(Payment.paid_date != None, Payment.paid_date >= start, Payment.paid_date < end),  # noqa: E711
                    db.and_(Payment.paid_date == None, Payment.due_date >= start, Payment.due_date < end),  # noqa: E711
                ),
            )
            .scalar()
            or 0
        )
        try:
            monthly_income.append(float(total))
        except Exception:
            monthly_income.append(0.0)

    monthly_expenses: list[float] = []
    for start, end in month_ranges:
        try:
            total = (
                db.session.query(db.func.coalesce(db.func.sum(Expense.amount), 0))
                .filter(Expense.spent_at >= start, Expense.spent_at < end)
                .scalar()
                or 0
            )
            monthly_expenses.append(float(total))
        except Exception:
            monthly_expenses.append(0.0)

    monthly_profit = [round((monthly_income[i] - monthly_expenses[i]), 2) for i in range(len(month_ranges))]

    # -----------------------
    # Additional KPIs & charts for dashboard cards
    # -----------------------

    # Properties/Units occupancy breakdown
    units_total = (
        (db.session.query(db.func.count(Apartment.id)).scalar() or 0)
        + (
            db.session.query(db.func.count(Property.id))
            .filter(Property.property_type == "apartment")
            .scalar()
            or 0
        )
    )
    units_available = (
        (db.session.query(db.func.count(Apartment.id)).filter(Apartment.status == "available").scalar() or 0)
        + (
            db.session.query(db.func.count(Property.id))
            .filter(Property.property_type == "apartment", Property.status == "available")
            .scalar()
            or 0
        )
    )
    units_occupied = (
        (db.session.query(db.func.count(Apartment.id)).filter(Apartment.status == "occupied").scalar() or 0)
        + (
            db.session.query(db.func.count(Property.id))
            .filter(Property.property_type == "apartment", Property.status == "occupied")
            .scalar()
            or 0
        )
    )

    # Maintenance and complaints counts
    maint_new = db.session.query(db.func.count(MaintenanceRequest.id)).filter(MaintenanceRequest.status == "new").scalar() or 0
    maint_in_progress = (
        db.session.query(db.func.count(MaintenanceRequest.id)).filter(MaintenanceRequest.status == "in_progress").scalar() or 0
    )
    maint_resolved = db.session.query(db.func.count(MaintenanceRequest.id)).filter(MaintenanceRequest.status == "resolved").scalar() or 0
    maint_closed = db.session.query(db.func.count(MaintenanceRequest.id)).filter(MaintenanceRequest.status == "closed").scalar() or 0

    comp_new = db.session.query(db.func.count(Complaint.id)).filter(Complaint.status == "new").scalar() or 0
    comp_reviewing = db.session.query(db.func.count(Complaint.id)).filter(Complaint.status == "reviewing").scalar() or 0
    comp_resolved = db.session.query(db.func.count(Complaint.id)).filter(Complaint.status == "resolved").scalar() or 0
    comp_closed = db.session.query(db.func.count(Complaint.id)).filter(Complaint.status == "closed").scalar() or 0

    # Contracts status counts
    contract_status_rows = (
        db.session.query(Contract.status, db.func.count(Contract.id))
        .group_by(Contract.status)
        .all()
    )
    contracts_by_status = {k or "unknown": int(v or 0) for k, v in contract_status_rows}
    contracts_total = sum(contracts_by_status.values())
    contracts_active = contracts_by_status.get("active", 0)

    # Payments counts
    paid_count = db.session.query(db.func.count(Payment.id)).filter(Payment.status == "paid").scalar() or 0
    overdue_count = (
        db.session.query(db.func.count(Payment.id))
        .filter(Payment.status != "paid", Payment.due_date < today)
        .scalar()
        or 0
    )
    upcoming_count = (
        db.session.query(db.func.count(Payment.id))
        .filter(Payment.status != "paid", Payment.due_date >= today, Payment.due_date <= (today + timedelta(days=14)))
        .scalar()
        or 0
    )

    # Recent items
    recent_properties = Property.query.order_by(Property.created_at.desc()).limit(5).all()
    recent_contracts = Contract.query.order_by(Contract.created_at.desc()).limit(5).all()
    recent_payments = Payment.query.order_by(Payment.created_at.desc()).limit(5).all()

    return render_template(
        "accountant/dashboard.html",
        payments=payments,
        total_paid=total_income,
        total_income=total_income,
        total_expenses=total_expenses,
        net_profit=net_profit,
        unpaid_count=unpaid_count,
        today=today,
        due_soon=due_soon,
        overdue=overdue,
        month_labels=month_labels,
        monthly_income=monthly_income,
        monthly_expenses=monthly_expenses,
        monthly_profit=monthly_profit,
        # occupancy chart
        occupancy_labels=[_("Available"), _("Occupied")],
        occupancy_values=[units_available, units_occupied],
        # kpis for cards
        units_total=units_total,
        units_available=units_available,
        units_occupied=units_occupied,
        contracts_total=contracts_total,
        contracts_active=contracts_active,
        contracts_by_status=contracts_by_status,
        paid_count=paid_count,
        overdue_count=overdue_count,
        upcoming_count=upcoming_count,
        maint_new=maint_new,
        maint_in_progress=maint_in_progress,
        maint_resolved=maint_resolved,
        maint_closed=maint_closed,
        comp_new=comp_new,
        comp_reviewing=comp_reviewing,
        comp_resolved=comp_resolved,
        comp_closed=comp_closed,
        # recents
        recent_properties=recent_properties,
        recent_contracts=recent_contracts,
        recent_payments=recent_payments,
    )


# -----------------------
# Accountant lists (read-only): Properties, Contracts, Maintenance, Complaints
# -----------------------


@accountant_bp.route("/properties")
@login_required
@accountant_required
def properties_list():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    ptype = (request.args.get("type") or "").strip()  # building | apartment | all

    props_q = Property.query
    if ptype in {"building", "apartment"}:
        props_q = props_q.filter(Property.property_type == ptype)
    if status in {"available", "occupied"}:
        props_q = props_q.filter(Property.status == status)
    if q:
        like = f"%{q}%"
        props_q = props_q.filter(Property.title.ilike(like))
    props = props_q.order_by(Property.created_at.desc()).all()

    # For quick occupancy totals on the list page
    total_buildings = Property.query.filter_by(property_type="building").count()
    total_standalone = Property.query.filter_by(property_type="apartment").count()

    return render_template(
        "accountant/properties.html",
        properties=props,
        total_buildings=total_buildings,
        total_standalone=total_standalone,
        q=q,
        selected_status=status or None,
        selected_type=ptype or "all",
    )


@accountant_bp.route("/contracts/list")
@login_required
@accountant_required
def contracts_list_accountant():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    contracts_q = Contract.query
    if status:
        contracts_q = contracts_q.filter(Contract.status == status)
    if start:
        try:
            from datetime import datetime as _dt

            start_d = _dt.strptime(start, "%Y-%m-%d").date()
            contracts_q = contracts_q.filter(Contract.start_date >= start_d)
        except Exception:
            pass
    if end:
        try:
            from datetime import datetime as _dt

            end_d = _dt.strptime(end, "%Y-%m-%d").date()
            contracts_q = contracts_q.filter(Contract.end_date <= end_d)
        except Exception:
            pass
    if q:
        like = f"%{q}%"
        from sqlalchemy import or_

        contracts_q = contracts_q.join(Property).join(User).filter(
            or_(Property.title.ilike(like), User.username.ilike(like))
        )
    contracts = contracts_q.order_by(Contract.created_at.desc()).all()

    return render_template("accountant/contracts.html", contracts=contracts, q=q, selected_status=status or None, start=start, end=end)


@accountant_bp.route("/maintenance")
@login_required
@accountant_required
def maintenance_list_accountant():
    status = (request.args.get("status") or "").strip()
    maint_q = MaintenanceRequest.query
    if status:
        maint_q = maint_q.filter(MaintenanceRequest.status == status)
    maintenance_requests = maint_q.order_by(MaintenanceRequest.created_at.desc()).all()
    return render_template("accountant/maintenance.html", maintenance_requests=maintenance_requests, selected_status=status or None)


@accountant_bp.route("/complaints")
@login_required
@accountant_required
def complaints_list_accountant():
    status = (request.args.get("status") or "").strip()
    complaints_q = Complaint.query
    if status:
        complaints_q = complaints_q.filter(Complaint.status == status)
    complaints = complaints_q.order_by(Complaint.created_at.desc()).all()
    return render_template("accountant/complaints.html", complaints=complaints, selected_status=status or None)


# -----------------------
# Tenants: list + detail/statement
# -----------------------


@accountant_bp.route("/tenants")
@login_required
@accountant_required
def tenants_list():
    """List all tenants with quick search."""
    q = (request.args.get("q") or "").strip()
    tenants_q = User.query.filter_by(role="tenant")
    if q:
        from sqlalchemy import or_

        like = f"%{q}%"
        tenants_q = tenants_q.filter(or_(User.username.ilike(like), User.phone.ilike(like)))
    tenants = tenants_q.order_by(User.created_at.desc()).all()
    return render_template("accountant/tenants_list.html", tenants=tenants, q=q)


@accountant_bp.route("/tenants/<int:tenant_id>", methods=["GET", "POST"])
@login_required
@accountant_required
def tenant_detail(tenant_id: int):
    """Show tenant statement and allow recording payments."""
    tenant = User.query.get_or_404(tenant_id)
    if tenant.role != "tenant":
        return abort(404)

    # Create a payment for this tenant (against a selected contract)
    if request.method == "POST":
        contract_id = request.form.get("contract_id", type=int)
        amount = request.form.get("amount", type=float)
        due_date = request.form.get("due_date")
        method = (request.form.get("method") or "").strip() or None
        status = (request.form.get("status") or "unpaid").strip().lower()

        contract = Contract.query.get(contract_id) if contract_id else None
        if not (contract and contract.tenant_id == tenant.id and amount and amount > 0 and due_date):
            flash(_("Invalid payment data"), "danger")
            return redirect(url_for("accountant.tenant_detail", tenant_id=tenant.id))

        p = Payment(
            contract_id=contract.id,
            amount=amount,
            due_date=due_date,
            method=method,
            status=("paid" if status == "paid" else "unpaid"),
        )
        # If marked as paid now, set paid_date and attempt journal posting later via mark toggle
        if p.status == "paid":
            from datetime import date as _date

            p.paid_date = _date.today()
        db.session.add(p)
        db.session.commit()
        # If paid, try to post cash receipt journal entry
        try:
            if p.status == "paid":
                _post_payment_cash_receipt(p)
        except Exception:
            pass
        flash(_("Payment recorded for tenant"), "success")
        return redirect(url_for("accountant.tenant_detail", tenant_id=tenant.id))

    # Data for statement
    contracts = Contract.query.filter_by(tenant_id=tenant.id).order_by(Contract.created_at.desc()).all()
    payments = (
        Payment.query.join(Contract, Payment.contract_id == Contract.id)
        .filter(Contract.tenant_id == tenant.id)
        .order_by(Payment.due_date.desc())
        .all()
    )

    total_amount = sum(float(p.amount or 0) for p in payments)
    total_paid_amount = sum(float(p.amount or 0) for p in payments if (p.status or "").lower() == "paid")
    total_unpaid_amount = round(total_amount - total_paid_amount, 2)

    return render_template(
        "accountant/tenant_detail.html",
        tenant=tenant,
        contracts=contracts,
        payments=payments,
        total_amount=round(total_amount, 2),
        total_paid_amount=round(total_paid_amount, 2),
        total_unpaid_amount=total_unpaid_amount,
    )


@accountant_bp.route("/payments/<int:payment_id>/mark", methods=["POST"])
@login_required
@accountant_required
def mark_payment(payment_id: int):
    payment = Payment.query.get_or_404(payment_id)
    new_status = request.form.get("status")
    if new_status in {"paid", "unpaid"}:
        payment.status = new_status
        db.session.commit()
        # Auto-post journal for payment status changes
        try:
            if new_status == "paid":
                _post_payment_cash_receipt(payment)
            else:
                _reverse_payment_cash_receipt(payment)
        except Exception:
            # Do not fail user flow on posting errors
            pass
        flash(_("Payment status updated"), "success")
    return redirect(url_for("accountant.dashboard"))


@accountant_bp.route("/payments/<int:payment_id>/invoice")
@login_required
@accountant_required
def generate_invoice(payment_id: int):
    payment = Payment.query.get_or_404(payment_id)
    # Create PDF in memory
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    c.setFont("Helvetica-Bold", 16)
    c.drawString(72, height - 72, "Invoice")
    c.setFont("Helvetica", 12)
    c.drawString(72, height - 110, f"Invoice for Payment ID: {payment.id}")
    c.drawString(72, height - 130, f"Contract ID: {payment.contract_id}")
    c.drawString(72, height - 150, f"Amount: {payment.amount}")
    c.drawString(72, height - 170, f"Due Date: {payment.due_date}")
    c.drawString(72, height - 190, f"Status: {payment.status}")
    c.showPage()
    c.save()
    buffer.seek(0)

    # Save to disk
    invoices_dir = os.path.join(current_app.config["UPLOAD_FOLDER"], "invoices")
    os.makedirs(invoices_dir, exist_ok=True)
    file_name = f"invoice_{payment.id}.pdf"
    file_path = os.path.join(invoices_dir, file_name)
    with open(file_path, "wb") as f:
        f.write(buffer.getvalue())

    # Create or update DB record
    if payment.invoice:
        payment.invoice.file_path = f"invoices/{file_name}"
    else:
        inv = Invoice(payment_id=payment.id, file_path=f"invoices/{file_name}")
        db.session.add(inv)
    db.session.commit()
    # Post AR and Rental Income for invoice (idempotent)
    try:
        _post_invoice_revenue(payment)
    except Exception:
        pass
    return redirect(url_for("accountant.dashboard"))


@accountant_bp.route("/invoices/<int:payment_id>/download")
@login_required
@accountant_required
def download_invoice(payment_id: int):
    payment = Payment.query.get_or_404(payment_id)
    if not payment.invoice:
        return abort(404)
    file_path = os.path.join(current_app.config["UPLOAD_FOLDER"], payment.invoice.file_path)
    return send_file(file_path, mimetype="application/pdf", as_attachment=True, download_name=os.path.basename(file_path))


@accountant_bp.route("/export/excel")
@login_required
@accountant_required
def export_payments_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    ws.append(["ID", "Contract", "Amount", "Due Date", "Paid Date", "Method", "Status"])
    for p in Payment.query.order_by(Payment.due_date.asc()).all():
        ws.append([p.id, p.contract_id, float(p.amount), str(p.due_date), str(p.paid_date or ""), p.method or "", p.status])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="payments.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@accountant_bp.route("/export/pdf")
@login_required
@accountant_required
def export_payments_pdf():
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    c.setFont("Helvetica-Bold", 16)
    c.drawString(72, height - 72, "Payments Report")
    c.setFont("Helvetica", 11)
    y = height - 110
    for p in Payment.query.order_by(Payment.due_date.asc()).all():
        line = f"#{p.id} Contract:{p.contract_id} Due:{p.due_date} Amount:{p.amount} Status:{p.status}"
        c.drawString(72, y, line)
        y -= 18
        if y < 72:
            c.showPage()
            y = height - 72
    c.showPage()
    c.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="payments.pdf", mimetype="application/pdf")


# -----------------------
# Invoices: list + exports
# -----------------------


@accountant_bp.route("/invoices")
@login_required
@accountant_required
def invoices_list():
    # Show payments with optional invoice attached; allow filter has_invoice
    has_invoice = request.args.get("has_invoice")
    q = Payment.query.order_by(Payment.due_date.desc())
    if has_invoice == "yes":
        q = q.join(Invoice, isouter=True).filter(Invoice.id != None)  # noqa: E711
    elif has_invoice == "no":
        q = q.join(Invoice, isouter=True).filter(Invoice.id == None)  # noqa: E711
    payments = q.all()
    return render_template("accountant/invoices.html", payments=payments, has_invoice=has_invoice)


@accountant_bp.route("/invoices/export.xlsx")
@login_required
@accountant_required
def export_invoices_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    ws.append(["PaymentID", "HasInvoice", "DueDate", "Amount", "Status", "InvoicePath"])
    rows = (
        db.session.query(Payment, Invoice)
        .join(Invoice, Payment.id == Invoice.payment_id, isouter=True)
        .order_by(Payment.due_date.desc())
        .all()
    )
    for p, inv in rows:
        ws.append([p.id, "yes" if inv else "no", str(p.due_date), float(p.amount), p.status, (inv.file_path if inv else "")])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="invoices.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@accountant_bp.route("/invoices/export.pdf")
@login_required
@accountant_required
def export_invoices_pdf():
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    c.setFont("Helvetica-Bold", 16)
    c.drawString(72, height - 72, "Invoices Report")
    c.setFont("Helvetica", 11)
    y = height - 110
    rows = (
        db.session.query(Payment, Invoice)
        .join(Invoice, Payment.id == Invoice.payment_id, isouter=True)
        .order_by(Payment.due_date.desc())
        .all()
    )
    for p, inv in rows:
        line = f"P#{p.id} Due:{p.due_date} Amount:{p.amount} Status:{p.status} HasInv:{'yes' if inv else 'no'}"
        c.drawString(72, y, line)
        y -= 18
        if y < 72:
            c.showPage()
            y = height - 72
    c.showPage()
    c.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="invoices.pdf", mimetype="application/pdf")


# -----------------------
# Payments: list + create
# -----------------------


@accountant_bp.route("/payments", methods=["GET", "POST"])
@login_required
@accountant_required
def payments_list():
    if request.method == "POST":
        contract_id = request.form.get("contract_id", type=int)
        amount = request.form.get("amount", type=float)
        due_date = request.form.get("due_date")
        method = (request.form.get("method") or "").strip() or None
        if not (contract_id and amount and amount > 0 and due_date):
            flash(_("Invalid payment data"), "danger")
            return redirect(url_for("accountant.payments_list"))
        if not Contract.query.get(contract_id):
            flash(_("Contract not found"), "danger")
            return redirect(url_for("accountant.payments_list"))
        p = Payment(contract_id=contract_id, amount=amount, due_date=due_date, method=method, status="unpaid")
        db.session.add(p)
        db.session.commit()
        flash(_("Payment created"), "success")
        return redirect(url_for("accountant.payments_list"))

    status = request.args.get("status")
    q = Payment.query
    if status in {"paid", "unpaid"}:
        q = q.filter(Payment.status == status)
    q = q.order_by(Payment.due_date.asc())
    payments = q.all()
    return render_template("accountant/payments.html", payments=payments, selected_status=status)


# -----------------------
# Expenses export (Excel/PDF)
# -----------------------


@accountant_bp.route("/expenses/export.xlsx")
@login_required
@accountant_required
def export_expenses_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    ws.append(["ID", "Date", "Description", "Category", "Vendor", "Amount"])
    for e in Expense.query.order_by(Expense.spent_at.desc()).all():
        ws.append([e.id, str(e.spent_at), e.description, e.category or "", e.vendor or "", float(e.amount)])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="expenses.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@accountant_bp.route("/expenses/export.pdf")
@login_required
@accountant_required
def export_expenses_pdf():
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    c.setFont("Helvetica-Bold", 16)
    c.drawString(72, height - 72, "Expenses Report")
    c.setFont("Helvetica", 11)
    y = height - 110
    for e in Expense.query.order_by(Expense.spent_at.desc()).all():
        line = f"#{e.id} {e.spent_at} {e.description} {e.category or ''} {e.vendor or ''} {e.amount}"
        c.drawString(72, y, line)
        y -= 18
        if y < 72:
            c.showPage()
            y = height - 72
    c.showPage()
    c.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="expenses.pdf", mimetype="application/pdf")


# -----------------------
# Unpaid report shortcut
# -----------------------


@accountant_bp.route("/reports/unpaid")
@login_required
@accountant_required
def report_unpaid():
    payments = (
        Payment.query.filter(Payment.status != "paid").order_by(Payment.due_date.asc()).all()
    )
    return render_template("accountant/payments.html", payments=payments, selected_status="unpaid")


@accountant_bp.route("/overview")
@login_required
@accountant_required
def financial_overview():
    """Comprehensive financial overview: totals and last 12 months breakdown."""
    # --- Totals ---
    total_income = (
        db.session.query(db.func.coalesce(db.func.sum(Payment.amount), 0))
        .filter_by(status="paid")
        .scalar()
        or 0
    )

    # Expenses may be in an optional "expenses" table
    total_expenses = 0
    monthly_expenses: list[float] = []
    expenses_date_column = None
    try:
        inspector = db.inspect(db.engine)
        if inspector.has_table("expenses"):
            # Sum total expenses
            res = db.session.execute(text("SELECT COALESCE(SUM(amount), 0) FROM expenses"))
            total_expenses = res.scalar() or 0
            # Try to detect a usable date column for monthly breakdown
            cols = inspector.get_columns("expenses")
            names = {c.get("name") for c in cols}
            for candidate in ("spent_at", "date", "created_at"):
                if candidate in names:
                    expenses_date_column = candidate
                    break
    except Exception:
        total_expenses = 0
        expenses_date_column = None

    # --- Build last 12 month ranges (ascending) ---
    def add_months(year: int, month: int, delta: int) -> tuple[int, int]:
        total = year * 12 + (month - 1) + delta
        y = total // 12
        m = (total % 12) + 1
        return y, m

    def first_of_month(y: int, m: int) -> date:
        return date(y, m, 1)

    def next_month_start(y: int, m: int) -> date:
        y2, m2 = add_months(y, m, 1)
        return date(y2, m2, 1)

    y0, m0 = date.today().year, date.today().month
    month_ranges: list[tuple[date, date]] = []
    month_labels: list[str] = []
    for k in range(11, -1, -1):
        yk, mk = add_months(y0, m0, -k)
        start = first_of_month(yk, mk)
        end = next_month_start(yk, mk)
        month_ranges.append((start, end))
        month_labels.append(f"{yk:04d}-{mk:02d}")

    # --- Monthly income ---
    monthly_income: list[float] = []
    for start, end in month_ranges:
        total = (
            db.session.query(db.func.coalesce(db.func.sum(Payment.amount), 0))
            .filter(
                Payment.status == "paid",
                db.or_(
                    db.and_(Payment.paid_date != None, Payment.paid_date >= start, Payment.paid_date < end),
                    db.and_(Payment.paid_date == None, Payment.due_date >= start, Payment.due_date < end),
                ),
            )
            .scalar()
            or 0
        )
        try:
            monthly_income.append(float(total))
        except Exception:
            monthly_income.append(0.0)

    # --- Monthly expenses (if table and date column exist) ---
    if expenses_date_column is not None:
        monthly_expenses = []
        for start, end in month_ranges:
            q = text(
                f"SELECT COALESCE(SUM(amount), 0) FROM expenses WHERE {expenses_date_column} >= :start AND {expenses_date_column} < :end"
            )
            try:
                r = db.session.execute(q, {"start": start, "end": end})
                monthly_expenses.append(float(r.scalar() or 0))
            except Exception:
                monthly_expenses.append(0.0)
    else:
        monthly_expenses = [0.0 for _ in month_ranges]

    # --- Profits ---
    try:
        profit_total = float(total_income) - float(total_expenses)
    except Exception:
        profit_total = 0.0

    monthly_profit = [round((monthly_income[i] - monthly_expenses[i]), 2) for i in range(len(month_ranges))]

    return render_template(
        "accountant/financial_overview.html",
        total_income=total_income,
        total_expenses=total_expenses,
        profit_total=profit_total,
        month_labels=month_labels,
        monthly_income=monthly_income,
        monthly_expenses=monthly_expenses,
        monthly_profit=monthly_profit,
    )


# -----------------------
# Chart of Accounts
# -----------------------


@accountant_bp.route("/accounts", methods=["GET", "POST"])
@login_required
@accountant_required
def accounts():
    if request.method == "POST":
        code = (request.form.get("code") or "").strip()
        name = (request.form.get("name") or "").strip()
        acc_type = (request.form.get("type") or "").strip()
        if code and name and acc_type in {"asset", "liability", "equity", "income", "expense"}:
            if not Account.query.filter_by(code=code).first():
                acc = Account(code=code, name=name, type=acc_type)
                db.session.add(acc)
                db.session.commit()
                flash(_("Account created"), "success")
            else:
                flash(_("Account code already exists"), "warning")
        else:
            flash(_("Invalid account data"), "danger")
        return redirect(url_for("accountant.accounts"))

    accounts = Account.query.order_by(Account.type.asc(), Account.code.asc()).all()
    return render_template("accountant/accounts.html", accounts=accounts)


# -----------------------
# Expenses CRUD (create + list)
# -----------------------


@accountant_bp.route("/expenses", methods=["GET", "POST"])
@login_required
@accountant_required
def expenses():
    if request.method == "POST":
        desc = (request.form.get("description") or "").strip()
        amount_raw = request.form.get("amount") or "0"
        category = (request.form.get("category") or "").strip()
        vendor = (request.form.get("vendor") or "").strip()
        spent_at = request.form.get("spent_at") or None
        try:
            amt = float(amount_raw)
        except Exception:
            amt = 0.0
        if desc and amt > 0:
            exp = Expense(description=desc, amount=amt, category=category or None, vendor=vendor or None, spent_at=spent_at)
            db.session.add(exp)
            db.session.commit()
            try:
                _post_expense_cash(exp)
            except Exception:
                pass
            flash(_("Expense recorded"), "success")
        else:
            flash(_("Invalid expense data"), "danger")
        return redirect(url_for("accountant.expenses"))

    expenses = Expense.query.order_by(Expense.spent_at.desc()).all()
    return render_template("accountant/expenses.html", expenses=expenses)


# -----------------------
# Journal Entry (simple two-line)
# -----------------------


@accountant_bp.route("/journal/new", methods=["GET", "POST"])
@login_required
@accountant_required
def journal_new():
    accounts = Account.query.order_by(Account.type.asc(), Account.code.asc()).all()
    if request.method == "POST":
        date_str = request.form.get("date")
        memo = request.form.get("memo")
        debit_account_id = request.form.get("debit_account_id")
        credit_account_id = request.form.get("credit_account_id")
        amount_raw = request.form.get("amount") or "0"
        try:
            amt = float(amount_raw)
        except Exception:
            amt = 0.0
        if not (debit_account_id and credit_account_id and amt > 0):
            flash(_("Invalid journal data"), "danger")
            return render_template("accountant/journal_new.html", accounts=accounts)
        je = JournalEntry(memo=memo or None)
        try:
            from datetime import datetime as _dt
            if date_str:
                je.date = _dt.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            pass
        db.session.add(je)
        db.session.flush()
        jl1 = JournalLine(entry_id=je.id, account_id=int(debit_account_id), debit=amt, credit=0)
        jl2 = JournalLine(entry_id=je.id, account_id=int(credit_account_id), debit=0, credit=amt)
        db.session.add_all([jl1, jl2])
        db.session.commit()
        flash(_("Journal entry created"), "success")
        return redirect(url_for("accountant.journal_new"))

    return render_template("accountant/journal_new.html", accounts=accounts)


# -----------------------
# General Ledger per account
# -----------------------


@accountant_bp.route("/ledger")
@login_required
@accountant_required
def ledger():
    account_id = request.args.get("account_id", type=int)
    accounts = Account.query.order_by(Account.type.asc(), Account.code.asc()).all()
    lines = []
    running = 0.0
    account = None
    if account_id:
        account = Account.query.get(account_id)
        if account:
            q = (
                db.session.query(JournalLine, JournalEntry)
                .join(JournalEntry, JournalLine.entry_id == JournalEntry.id)
                .filter(JournalLine.account_id == account_id)
                .order_by(JournalEntry.date.asc(), JournalEntry.id.asc(), JournalLine.id.asc())
            )
            for jl, je in q.all():
                debit = float(jl.debit or 0)
                credit = float(jl.credit or 0)
                if account.type in {"asset", "expense"}:
                    running += debit - credit
                else:
                    running += credit - debit
                lines.append({
                    "date": je.date,
                    "memo": je.memo,
                    "debit": debit,
                    "credit": credit,
                    "balance": round(running, 2),
                })
    return render_template("accountant/ledger.html", accounts=accounts, lines=lines, selected_account=account)


# -----------------------
# Trial Balance
# -----------------------


@accountant_bp.route("/trial-balance")
@login_required
@accountant_required
def trial_balance():
    rows = (
        db.session.query(
            JournalLine.account_id,
            db.func.coalesce(db.func.sum(JournalLine.debit), 0).label("debits"),
            db.func.coalesce(db.func.sum(JournalLine.credit), 0).label("credits"),
        )
        .group_by(JournalLine.account_id)
        .all()
    )
    data = []
    total_debits = 0.0
    total_credits = 0.0
    for account_id, debits, credits in rows:
        acc = Account.query.get(account_id)
        d = float(debits or 0)
        c = float(credits or 0)
        balance = d - c if acc.type in {"asset", "expense"} else c - d
        data.append({"account": acc, "debits": d, "credits": c, "balance": round(balance, 2)})
        total_debits += d
        total_credits += c
    return render_template("accountant/trial_balance.html", rows=data, total_debits=round(total_debits, 2), total_credits=round(total_credits, 2))


# -----------------------
# Income Statement
# -----------------------


@accountant_bp.route("/income-statement")
@login_required
@accountant_required
def income_statement():
    # Income accounts: credit-normal
    income_rows = (
        db.session.query(
            Account,
            db.func.coalesce(db.func.sum(JournalLine.debit), 0),
            db.func.coalesce(db.func.sum(JournalLine.credit), 0),
        )
        .join(JournalLine, JournalLine.account_id == Account.id)
        .filter(Account.type == "income")
        .group_by(Account.id)
        .all()
    )
    expense_rows = (
        db.session.query(
            Account,
            db.func.coalesce(db.func.sum(JournalLine.debit), 0),
            db.func.coalesce(db.func.sum(JournalLine.credit), 0),
        )
        .join(JournalLine, JournalLine.account_id == Account.id)
        .filter(Account.type == "expense")
        .group_by(Account.id)
        .all()
    )
    income_total = sum(float(c) - float(d) for _, d, c in income_rows)
    expense_total = sum(float(d) - float(c) for _, d, c in expense_rows)
    net_income = round(income_total - expense_total, 2)
    return render_template(
        "accountant/income_statement.html",
        income_rows=income_rows,
        expense_rows=expense_rows,
        income_total=round(income_total, 2),
        expense_total=round(expense_total, 2),
        net_income=net_income,
    )


# -----------------------
# Balance Sheet (as of today)
# -----------------------


@accountant_bp.route("/balance-sheet")
@login_required
@accountant_required
def balance_sheet():
    # Aggregate balances by account type
    def account_balance(acc: Account) -> float:
        sums = (
            db.session.query(
                db.func.coalesce(db.func.sum(JournalLine.debit), 0),
                db.func.coalesce(db.func.sum(JournalLine.credit), 0),
            )
            .filter(JournalLine.account_id == acc.id)
            .first()
        )
        d, c = float(sums[0] or 0), float(sums[1] or 0)
        return (d - c) if acc.type in {"asset", "expense"} else (c - d)

    assets = Account.query.filter_by(type="asset").order_by(Account.code).all()
    liabilities = Account.query.filter_by(type="liability").order_by(Account.code).all()
    equity = Account.query.filter_by(type="equity").order_by(Account.code).all()

    assets_rows = [(acc, round(account_balance(acc), 2)) for acc in assets]
    liabilities_rows = [(acc, round(account_balance(acc), 2)) for acc in liabilities]
    equity_rows = [(acc, round(account_balance(acc), 2)) for acc in equity]

    assets_total = round(sum(v for _, v in assets_rows), 2)
    liabilities_total = round(sum(v for _, v in liabilities_rows), 2)
    equity_total = round(sum(v for _, v in equity_rows), 2)
    return render_template(
        "accountant/balance_sheet.html",
        assets_rows=assets_rows,
        liabilities_rows=liabilities_rows,
        equity_rows=equity_rows,
        assets_total=assets_total,
        liabilities_total=liabilities_total,
        equity_total=equity_total,
    )


# -----------------------
# Accounts Receivable Aging
# -----------------------


@accountant_bp.route("/ar-aging")
@login_required
@accountant_required
def ar_aging():
    from datetime import date as _date
    today = _date.today()
    # Unpaid tenant payments
    payments = (
        db.session.query(Payment)
        .filter(Payment.status != "paid")
        .order_by(Payment.due_date.asc())
        .all()
    )
    buckets = {"current": 0.0, "1-30": 0.0, "31-60": 0.0, "61-90": 0.0, ">90": 0.0}
    rows = []
    for p in payments:
        days = (today - (p.due_date or today)).days
        amt = float(p.amount or 0)
        if days <= 0:
            buckets["current"] += amt
            bucket = "current"
        elif days <= 30:
            buckets["1-30"] += amt
            bucket = "1-30"
        elif days <= 60:
            buckets["31-60"] += amt
            bucket = "31-60"
        elif days <= 90:
            buckets["61-90"] += amt
            bucket = "61-90"
        else:
            buckets[">90"] += amt
            bucket = ">90"
        tenant_name = getattr(getattr(p.contract, "tenant", None), "username", "-")
        rows.append({
            "payment": p,
            "tenant": tenant_name,
            "days": days,
            "bucket": bucket,
            "amount": amt,
        })
    return render_template("accountant/ar_aging.html", rows=rows, buckets=buckets)


# -----------------------
# Helpers for Posting
# -----------------------


def _get_or_create_account(code: str, name: str, acc_type: str) -> Account:
    acc = Account.query.filter_by(code=code).first()
    if not acc:
        acc = Account(code=code, name=name, type=acc_type)
        db.session.add(acc)
        db.session.commit()
    return acc


def _default_accounts() -> dict:
    return {
        "cash": _get_or_create_account("1000", "Cash", "asset"),
        "ar": _get_or_create_account("1100", "Accounts Receivable", "asset"),
        "rent_income": _get_or_create_account("4000", "Rental Income", "income"),
        "expense_generic": _get_or_create_account("5000", "General Expenses", "expense"),
    }


def _post_invoice_revenue(payment: Payment) -> None:
    # If already posted for this payment as invoice, skip
    exists = JournalEntry.query.filter_by(source="invoice", source_id=payment.id).first()
    if exists:
        return
    acc = _default_accounts()
    je = JournalEntry(date=payment.due_date or date.today(), memo=f"Invoice for payment #{payment.id}", source="invoice", source_id=payment.id)
    db.session.add(je)
    db.session.flush()
    amount = float(payment.amount or 0)
    db.session.add_all([
        JournalLine(entry_id=je.id, account_id=acc["ar"].id, debit=amount, credit=0),
        JournalLine(entry_id=je.id, account_id=acc["rent_income"].id, debit=0, credit=amount),
    ])
    db.session.commit()


def _post_payment_cash_receipt(payment: Payment) -> None:
    # If already posted for this payment as cash receipt, skip
    exists = JournalEntry.query.filter_by(source="payment", source_id=payment.id).first()
    if exists:
        return
    acc = _default_accounts()
    je = JournalEntry(date=payment.paid_date or date.today(), memo=f"Cash receipt for payment #{payment.id}", source="payment", source_id=payment.id)
    db.session.add(je)
    db.session.flush()
    amount = float(payment.amount or 0)
    db.session.add_all([
        JournalLine(entry_id=je.id, account_id=acc["cash"].id, debit=amount, credit=0),
        JournalLine(entry_id=je.id, account_id=acc["ar"].id, debit=0, credit=amount),
    ])
    db.session.commit()


def _reverse_payment_cash_receipt(payment: Payment) -> None:
    exists = JournalEntry.query.filter_by(source="payment", source_id=payment.id).first()
    if not exists:
        return
    acc = _default_accounts()
    je = JournalEntry(date=date.today(), memo=f"Reversal cash receipt for payment #{payment.id}", source="payment_reverse", source_id=payment.id)
    db.session.add(je)
    db.session.flush()
    amount = float(payment.amount or 0)
    db.session.add_all([
        JournalLine(entry_id=je.id, account_id=acc["cash"].id, debit=0, credit=amount),
        JournalLine(entry_id=je.id, account_id=acc["ar"].id, debit=amount, credit=0),
    ])
    db.session.commit()


def _post_expense_cash(exp: Expense) -> None:
    acc = _default_accounts()
    je = JournalEntry(date=exp.spent_at or date.today(), memo=f"Expense: {exp.description}", source="expense", source_id=exp.id)
    db.session.add(je)
    db.session.flush()
    amount = float(exp.amount or 0)
    db.session.add_all([
        JournalLine(entry_id=je.id, account_id=acc["expense_generic"].id, debit=amount, credit=0),
        JournalLine(entry_id=je.id, account_id=acc["cash"].id, debit=0, credit=amount),
    ])
    db.session.commit()
